from openpyxl import load_workbook
from openpyxl import Workbook

def sheet_filter(sheet_ranges, ws):
    mark = False
    # 第0列(row[0])是Sti1.ACC
    print(sheet_ranges.rows)
    for row in sheet_ranges.rows:
        if mark == True and 1 == row[0].value:
            row[0].value = 2
            mark = False
        if 0 == row[0].value:
            mark = True
        ws.append(row)
    return ws

filename = 'filter0420'
wb = load_workbook('.'.join([filename,'xlsx']))
output = Workbook(write_only=True)

for sheet_name in wb.get_sheet_names():
    sheet_ranges = wb[sheet_name]
    ws = output.create_sheet(sheet_name)
    sheet_filter(sheet_ranges, ws)

output.save('.'.join([filename + '_output','xlsx']))
