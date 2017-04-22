from openpyxl import load_workbook
from openpyxl import Workbook

def sheet_filter(sheet_ranges, ws):
    mark = False

    for row in sheet_ranges.rows:
        if mark == True and 1 == row[1].value:
            #ws.append([row[2], ])
            row[1].value = 2
            mark = False
        if 0 == row[1].value:
            mark = True
        ws.append(row)
    return ws

filename = 'filter'
wb = load_workbook('.'.join([filename,'xlsx']))
output = Workbook(write_only=True)

for sheet_name in wb.get_sheet_names():
    sheet_ranges = wb[sheet_name]
    ws = output.create_sheet(sheet_name)
    sheet_filter(sheet_ranges, ws)

output.save('filter_output.xlsx')
